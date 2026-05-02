from fastapi import FastAPI, UploadFile, File
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pytesseract
import cv2
import numpy as np
import re
import os

app = FastAPI()

EXCEL_FILE = "pagos/pagos.xlsx"

@app.get("/")
def inicio():
    return {"mensaje": "API funcionando"}

def crear_excel():
    os.makedirs("pagos", exist_ok=True)

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append([
            "Fecha",
            "Hora",
            "Nombre",
            "Monto",
            "Tipo",
            "Operacion",
            "Estado",
            "Registrado por"
        ])
        wb.save(EXCEL_FILE)

def preprocesar_imagen(image_bytes):
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
    return thresh

def extraer_datos(texto):
    texto_limpio = texto.replace("\n", " ")

    monto = re.search(r"S/\.?\s*(\d+(?:[.,]\d{1,2})?)", texto_limpio, re.IGNORECASE)
    operacion = re.search(r"(?:Operaci[oó]n|Nro\.?\s*de\s*operaci[oó]n)[:\s]*(\d+)", texto_limpio, re.IGNORECASE)
    fecha = re.search(r"(\d{1,2}\s+\w+\.?\s+\d{4}|\d{2}/\d{2}/\d{4})", texto_limpio, re.IGNORECASE)

    tipo = "Yape" if "yape" in texto.lower() or "yapeaste" in texto.lower() else "Plin" if "plin" in texto.lower() else "Desconocido"

    monto_valor = monto.group(1).replace(",", ".") if monto else None

    return {
        "fecha": fecha.group(1) if fecha else datetime.now().strftime("%d/%m/%Y"),
        "hora": datetime.now().strftime("%H:%M:%S"),
        "nombre": "No identificado",
        "monto": monto_valor,
        "tipo": tipo,
        "operacion": operacion.group(1) if operacion else "No detectada",
        "estado": "Registrado" if monto_valor else "No válido",
        "registrado_por": "Bot automático",
        "valido": monto_valor is not None,
        "texto_raw": texto
    }

def registrar_pago(datos):
    crear_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Pagos"]

    # Si ya existe una fila de TOTAL, la borramos
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=1).value == "TOTAL RECAUDADO":
            ws.delete_rows(row)

    # Agregar nuevo pago
    ws.append([
        datos["fecha"],
        datos["hora"],
        datos["nombre"],
        datos["monto"],
        datos["tipo"],
        datos["operacion"],
        datos["estado"],
        datos["registrado_por"]
    ])

    # Calcular total
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        monto = row[3]
        estado = row[6]

        if estado == "Registrado" and monto is not None:
            total += float(monto)

    # Dejar una fila vacía
    ws.append([])

    # Escribir total abajo
    ws.append([
        "TOTAL RECAUDADO",
        "",
        "",
        total,
        "",
        "",
        "",
        ""
    ])

    wb.save(EXCEL_FILE)

@app.post("/procesar-imagen")
async def procesar_imagen(file: UploadFile = File(...)):
    contenido = await file.read()
    img = preprocesar_imagen(contenido)
    texto = pytesseract.image_to_string(img, lang="spa")
    datos = extraer_datos(texto)

    if datos["valido"]:
        registrar_pago(datos)

    return datos

@app.get("/reporte")
def reporte():
    crear_excel()
    hoy = datetime.now().strftime("%d/%m/%Y")

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Pagos"]

    total = 0
    cantidad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = row[0]
        monto = row[3]
        estado = row[6]

        if fecha == hoy and estado == "Registrado":
            total += float(monto)
            cantidad += 1

    return {
        "fecha": hoy,
        "total_recaudado": total,
        "cantidad_pagos": cantidad
    }